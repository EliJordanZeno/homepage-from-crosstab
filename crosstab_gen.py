'''
author: Eli Jordan @ZenoGroup
date: 7/25/2023
This script generates a formatted 'Homepage' and 'Crosstab' sheet from the downloaded crosstab data from Mercury.
Make sure to check the input file before running this script. (delete screener and demo Questions)
garbage in —— garbage out
'''


import pandas as pd 
import numpy as np 
import openpyxl as pxl
from copy_xl import copy_sheet
import re

_DEBUG = False #used to print data in use 
#set to true for Verbose output 

class Question: 
    ''' 
    Question Class representing each survey question 
    methods - none 
    properties - number_internal, matrix_options, number, question_text, matrix_option, base 
    ''' 
    def __init__(self,raw): 
        #parse raw into standard parameters 
        main = raw.split(":",1) # if split is more than 1, questions with colons will not be captured
        if _DEBUG:
            print(main)
        self.number_internal = main[0] 
        self.matrix_options = [] 

        q = main[1].split("_") 

        #length of 2 should be a single part question 

        if len(q) == 2: 
            self.number = q[0][1:] 
            self.question_text = q[1] 
            self.matrix_option = None 
            self.base = True 
        #length of 4 is Q#_#_QuestionText_Option 

        elif len(q) == 4: #should only ever be 4 
            self.number = q[0][1:] 
            self.question_text = q[2] 
            self.matrix_option = q[3] 
            self.matrix_options.append(self.matrix_option) 
            self.base = False 

        #other strings that fit the format but aren't questions 
        else: 
            self.number = "000" 
            self.question_text = "000" 
            self.base = None 
        #print("parsed",q) 


class Survey: 
    """ 

    Class instantiated by an Excel file in the format from crosstabs 

    Methods: 

    output_to_file(file_name) - cleans questions from input xlsx file and outputs to file_name 

    Properties:

    obj_dict - the dictionary that organizes and stores all of the questions in the crosstab file

    """ 

    def __init__(self,file_name): 
        regex = r"\w{3,4}:\sQ|\w{3}:\sADD|\w{3}:\sP" #finds all strings with Question format
        df = pd.read_excel(file_name, sheet_name='Crosstab').iloc[:,0]
        wb = pxl.load_workbook(filename=file_name)
        sheet = wb.active
        # for cell in sheet["A"]:
        #     if bool(re.search(regex,cell.value)):
        #         selected_cells.append(cell.value)
        selected_cells = df[df.str.contains(regex, regex=True, na=False,case=True)] 

        questions_list = selected_cells.to_list() 

        if _DEBUG:
            pd.options.display.max_rows = 999
            print(df)
            #print(selected_cells)
        self.obj_dict = {} #dictionary storing questions by number 

        for q in questions_list: 
            new = Question(q) 
            if new.number == "000" or new.number == "QC" or new.number == "PRG": 
                continue 
            if new.number not in self.obj_dict.keys(): 
                self.obj_dict[new.number] = new 
                continue 
            else: 
                self.obj_dict[new.number].matrix_options.append(new.matrix_option) 

        if _DEBUG: 
            for i in self.obj_dict.keys(): 
                print(i,self.obj_dict[i].question_text,self.obj_dict[i].matrix_options) 

    def output_to_file(self,file_name): 
        #obj_dict -> df -> excel 
        #this should be fun to turn into a human-readable format 
        df_dict = {} 
        for k,i in self.obj_dict.items(): 
            df_dict[i.number+ ": " + i.question_text] = i.matrix_options 
        q_list = [] 
        o_list = [] 
        for question, options in df_dict.items(): 
            if len(df_dict[question]) == 1: 
                q_list.append(question) 
                o_list.append('') 
            else: 
                q_list.append(question) 
                o_list.append('')
                option_counter = 0 
                for value in options: 
                    option_counter+=1
                    o_list.append(f'Option {option_counter}: ' + value) 
                    q_list.append('') 
        final = pd.DataFrame(list(zip(q_list, o_list)), columns = ['Questions', 'Options']) 

        if _DEBUG: 
            print(final) 
        writer = pd.ExcelWriter(file_name)
        final.to_excel(writer,sheet_name="Homepage",index=False) 
        writer._save()

def create_links(workbook:pxl.Workbook):
    regex = r"\w{3,4}:\sQ|\w{3}:\sADD|\w{3}:\sP" # for finding formatted Questions
    #finding cells with Q or O and adding them to a dict with the form:
    # {
    # Q##O: CellObject
    # }
    # do this for both homepage and crosstab
    # iterate over the keys of dict (either one they should be the same) and add link to home page cell

    home_sheet = workbook["Homepage"]
    crosstab_sheet = workbook["Crosstab"]

    questions = home_sheet['A:B']
    question_data = crosstab_sheet["A"]
    homepage_dict = {}
    for column in questions:
        print(column)
        for i,cell in enumerate(column,1):
            if cell.value:
                if column[i+1].value: # no options
                    print(cell.value)
                    pass
                else: #definitely has options
                    pass
                work_arr = cell.value.split(":",1)
                homepage_dict[work_arr[0]] = cell
    print(homepage_dict)

if __name__ == "__main__":

    source_file_path = "KHC_COUNTRIES_RAW.xlsx" # input file (from Mercury), after checking typos and formatting
    output_file_path = "./KHC_COUNTRIES_CROSSTAB.xlsx" # output file

    #generates homepage cells
    s = Survey(source_file_path) 
    s.output_to_file(output_file_path)

    #copies data from source file to output file
    outputWB = pxl.load_workbook(filename=output_file_path)
    output_sheet = outputWB.create_sheet("Crosstab")

    inputWB = pxl.load_workbook(filename=source_file_path)
    input_sheet = inputWB["Crosstab"]

    copy_sheet(input_sheet,output_sheet)

    #create_links(outputWB)
    # link creation, and question formatting goes here
    outputWB.save(output_file_path)